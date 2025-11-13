import json
from typing import List, Dict, Any
from datetime import datetime

class ScheduleParser:
    """Парсер расписания из Excel с сохранением в файл"""
    
    TYPE_MAPPING = {
        'лекция': 'lecture',
        'практика': 'practice', 
        'лабораторная': 'lab',
        'лаба': 'lab',
        'семинар': 'seminar',
        'консультация': 'consultation',
        'Лекция': 'lecture',
        'Практика': 'practice',
        'Лабораторная': 'lab',
        'Семинар': 'seminar'
    }
    
    DAYS_ORDER = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']
    
    def __init__(self, university_id: int):
        self.university_id = university_id
    
    def parse_excel(self, file_path: str) -> List[Dict[str, Any]]:
        """Парсит Excel файл"""
        print(f"Читаем файл: {file_path}")
        
        try:
            try:
                from openpyxl import load_workbook
            except ImportError:
                print("openpyxl не установлен. Установите: pip install openpyxl")
                return []
            
            workbook = load_workbook(file_path)
            sheet = workbook.active
            
            lessons = []
            
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value if cell.value else "")
            
            print(f"📋 Заголовки: {headers}")
            
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not row[0]:
                    continue
                row_data = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        row_data[header] = str(row[i]) if row[i] is not None else ""
                    else:
                        row_data[header] = ""
                
                if not row_data.get('День') or not row_data.get('Предмет'):
                    continue
                
                lesson = {
                    'day': row_data['День'].strip(),
                    'time': row_data['Время'].strip(),
                    'subject': row_data['Предмет'].strip(),
                    'type': self._normalize_type(row_data.get('Тип', 'Практика')),
                    'room': row_data.get('Аудитория', '').strip(),
                    'teacher': row_data.get('Преподаватель', '').strip(),
                    'group': row_data.get('группа', '').strip()
                }
                
                lessons.append(lesson)
                print(f"Строка {row_num}: {lesson['day']} - {lesson['subject']}")
            
            print(f"Прочитано занятий: {len(lessons)}")
            return self._group_by_days(lessons)
            
        except FileNotFoundError:
            print(f"Файл не найден: {file_path}")
            return []
        except Exception as e:
            print(f"Ошибка при чтении файла: {e}")
            return []
    
    def _normalize_type(self, lesson_type: str) -> str:
        if not lesson_type:
            return 'practice'
        return self.TYPE_MAPPING.get(lesson_type, lesson_type.lower())
    
    def _group_by_days(self, lessons: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        days_dict = {}
        for lesson in lessons:
            day = lesson['day']
            if day not in days_dict:
                days_dict[day] = []
            lesson_obj = {
                'time': lesson['time'],
                'subject': lesson['subject'],
                'type': lesson['type'],
                'room': lesson['room'],
                'teacher': lesson['teacher']
            }
            days_dict[day].append(lesson_obj)
        
        result = []
        for day in self.DAYS_ORDER:
            if day in days_dict:
                result.append({
                    'university_id': self.university_id,
                    'day': day,
                    'lessons': days_dict[day]
                })
        
        return result
    
    def save_to_js_file(self, schedule_data: List[Dict[str, Any]], output_file: str = "schedule_output.js") -> str:
        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write("// Автоматически сгенерированное расписание\n")
                f.write("// Дата создания: " + datetime.now().strftime("%d.%m.%Y %H:%M") + "\n\n")
                f.write("schedule: [\n")
                
                for i, day_schedule in enumerate(schedule_data):
                    f.write("    {\n")
                    f.write(f'        university_id: {day_schedule["university_id"]},\n')
                    f.write(f'        day: "{day_schedule["day"]}",\n')
                    f.write("        lessons: [\n")
                    
                    for j, lesson in enumerate(day_schedule["lessons"]):
                        comma = "," if j < len(day_schedule["lessons"]) - 1 else ""
                        f.write(f'            {{ time: "{lesson["time"]}", subject: "{lesson["subject"]}", type: "{lesson["type"]}", room: "{lesson["room"]}", teacher: "{lesson["teacher"]}" }}{comma}\n')
                    
                    f.write("        ]\n")
                    f.write(f"    }}{',' if i < len(schedule_data) - 1 else ''}\n")
                    f.write("\n")
                
                f.write("]")
            
            print(f"Расписание сохранено в файл: {output_file}")
            return output_file
            
        except Exception as e:
            print(f"Ошибка при сохранении: {e}")
            return ""


def main():
    """Основная функция"""
    print("Парсер расписания - сохранение в файлы")
        
    try:
        university_id = int(input("Введите ID университета: "))
    except ValueError:
        print("Ошибка: ID университета должен быть числом!")
        return
    
    input_file = "get_shedule_example.xlsx"
    
    try:
        with open(input_file, 'r', encoding='utf-8'):
            pass
    except FileNotFoundError:
        print(f"Файл {input_file} не найден!")
        print("Убедитесь, что файл находится в той же папке, что и скрипт")
        return
    
    parser = ScheduleParser(university_id=university_id)
    schedule = parser.parse_excel(input_file)
    
    if schedule:
        js_file = parser.save_to_js_file(schedule, "new_schedule.js")
        total_lessons = sum(len(day['lessons']) for day in schedule)
        print(f"\nИтог:")
        print(f"- Университет: {parser.university_id}")
        print(f"- Дней с занятиями: {len(schedule)}")
        print(f"- Всего занятий: {total_lessons}")
        print(f"\nСоздан файл: {js_file}")
        
    else:
        print("\nНе удалось распарсить расписание")
        print("Проверьте структуру файла get_shedule_example.xlsx")

if __name__ == "__main__":
    main()